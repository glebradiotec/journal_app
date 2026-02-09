"""
Telegram-бот для Radiotec Journal App.
Слушает команды и отправляет бэкапы / Excel-выгрузки.
Запускается как systemd-сервис (journal-bot.service).
"""

import os
import time
import sqlite3
import requests
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TELEGRAM_BOT_TOKEN = '8568162243:AAFIJGHdgjb4swYCUuBU2pzMHggp9pRGMhA'
TELEGRAM_CHAT_ID = '134711555'
BOT_PASSWORD = '1845'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_APP_URL = 'http://127.0.0.1:8001'

API = f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}'
POLL_INTERVAL = 2  # секунды

# Авторизованные пользователи (chat_id -> True). Владелец всегда авторизован.
_authorized_users = {TELEGRAM_CHAT_ID: True}


def send_message(chat_id, text, reply_markup=None):
    """Отправить текстовое сообщение."""
    data = {'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'}
    if reply_markup:
        import json
        data['reply_markup'] = json.dumps(reply_markup)
    requests.post(f'{API}/sendMessage', data=data, timeout=15)


def send_document(chat_id, filepath, caption=''):
    """Отправить файл."""
    with open(filepath, 'rb') as f:
        requests.post(f'{API}/sendDocument', data={
            'chat_id': chat_id,
            'caption': caption,
        }, files={'document': (os.path.basename(filepath), f)}, timeout=60)


def get_main_keyboard():
    """Клавиатура с основными кнопками."""
    return {
        'keyboard': [
            [{'text': '📊 Excel-выгрузка'}, {'text': '💾 Бэкап БД'}],
            [{'text': '📈 Статистика'}],
        ],
        'resize_keyboard': True,
    }


def handle_start(chat_id):
    """Обработка /start."""
    send_message(chat_id,
        '👋 <b>Radiotec Journal App</b>\n\n'
        '📦 Каждый день в 3:00 сюда приходит бэкап БД и Excel-выгрузка.\n\n'
        'Доступные действия:',
        reply_markup=get_main_keyboard()
    )


def generate_excel(filepath):
    """Генерирует Excel-выгрузку статей напрямую из SQLite."""
    db_path = os.path.join(BASE_DIR, 'instance', 'journal.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    rows = c.execute('''
        SELECT a.id, a.title, a.authors, a.submission_date,
               a.payment_received, a.has_review, a.edited,
               j.name as journal_name, i.number as issue_num, i.year as issue_year
        FROM article a
        JOIN issue i ON a.issue_id = i.id
        JOIN journal j ON i.journal_id = j.id
        ORDER BY a.id DESC
    ''').fetchall()

    # Авторы из отдельной таблицы
    authors_map = {}
    for row in c.execute('SELECT article_id, full_name FROM article_author ORDER BY "order"'):
        authors_map.setdefault(row[0], []).append(row[1])
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Статьи"

    # Стили
    thin_border = Border(
        left=Side(style='thin', color='D0D7DE'), right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'), bottom=Side(style='thin', color='D0D7DE'),
    )
    title_font = Font(bold=True, size=14, color='24292F')
    subtitle_font = Font(size=11, color='57606A')
    header_fill = PatternFill(start_color='2D333B', end_color='2D333B', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_even_fill = PatternFill(start_color='F6F8FA', end_color='F6F8FA', fill_type='solid')
    status_yes_fill = PatternFill(start_color='DAFBE1', end_color='DAFBE1', fill_type='solid')
    status_yes_font = Font(bold=True, color='1A7F37', size=11)
    status_no_fill = PatternFill(start_color='FFEBE9', end_color='FFEBE9', fill_type='solid')
    status_no_font = Font(bold=True, color='CF222E', size=11)
    summary_fill = PatternFill(start_color='DDF4FF', end_color='DDF4FF', fill_type='solid')
    summary_font = Font(bold=True, size=11, color='0969DA')

    # Заголовок
    ws.merge_cells('A1:I1')
    ws['A1'].value = 'Экспорт статей — Radiotec Journal App'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    ws['A2'].value = f'Все статьи  |  Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(vertical='center')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    # Заголовки таблицы
    headers = ['№', 'Название', 'Авторы', 'Журнал', 'Выпуск', 'Дата поступления', 'Оплата', 'Рецензия', 'Редакт.']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 26

    # Данные
    paid_count = reviewed_count = edited_count = 0
    for row_idx, article in enumerate(rows, 1):
        r = row_idx + 4

        authors = ', '.join(authors_map.get(article['id'], [])) or article['authors'] or '-'
        journal_name = article['journal_name'] or '-'
        issue_info = f"№{article['issue_num']}/{article['issue_year']}" if article['issue_num'] else '-'

        is_paid = bool(article['payment_received'])
        is_reviewed = bool(article['has_review'])
        is_edited = bool(article['edited'])
        if is_paid: paid_count += 1
        if is_reviewed: reviewed_count += 1
        if is_edited: edited_count += 1

        row_data = [
            row_idx, article['title'] or '-', authors, journal_name,
            issue_info, article['submission_date'] or '-',
            'Да' if is_paid else 'Нет',
            'Да' if is_reviewed else 'Нет',
            'Да' if is_edited else 'Нет',
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 2))
            if row_idx % 2 == 0:
                cell.fill = row_even_fill

        for col in (7, 8, 9):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value == 'Да':
                cell.fill = status_yes_fill
                cell.font = status_yes_font
            else:
                cell.fill = status_no_fill
                cell.font = status_no_font

    total = len(rows)
    if total > 0:
        sr = total + 5
        ws.row_dimensions[sr].height = 26
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=2)
        sc = ws.cell(row=sr, column=1, value=f'Итого: {total} статей')
        sc.font = summary_font; sc.fill = summary_fill; sc.alignment = Alignment(vertical='center'); sc.border = thin_border
        ws.cell(row=sr, column=2).fill = summary_fill; ws.cell(row=sr, column=2).border = thin_border
        for c in range(3, 7):
            cell = ws.cell(row=sr, column=c, value='')
            cell.fill = summary_fill; cell.border = thin_border
        for col, count in [(7, paid_count), (8, reviewed_count), (9, edited_count)]:
            cell = ws.cell(row=sr, column=col, value=f'{count}/{total}')
            cell.font = summary_font; cell.fill = summary_fill
            cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin_border

    col_widths = {'A': 5, 'B': 42, 'C': 32, 'D': 22, 'E': 13, 'F': 17, 'G': 11, 'H': 11, 'I': 11}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = 'A5'

    wb.save(filepath)
    return total


def handle_export(chat_id):
    """Генерирует и отправляет Excel-выгрузку."""
    send_message(chat_id, '⏳ Генерирую Excel-выгрузку...')
    try:
        filename = f'articles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filepath = os.path.join(BASE_DIR, 'backups', filename)
        os.makedirs(os.path.join(BASE_DIR, 'backups'), exist_ok=True)
        total = generate_excel(filepath)
        send_document(chat_id, filepath,
            f'📊 Excel-выгрузка статей ({total} шт.)\n📅 {datetime.now().strftime("%d.%m.%Y %H:%M")}')
        os.remove(filepath)
    except Exception as e:
        send_message(chat_id, f'❌ Ошибка: {e}')


def handle_backup(chat_id):
    """Создать бэкап БД и отправить."""
    send_message(chat_id, '⏳ Создаю бэкап БД...')
    try:
        from backup import create_backup
        create_backup()

        # Найти последний бэкап
        backups_dir = os.path.join(BASE_DIR, 'backups')
        files = sorted(
            [f for f in os.listdir(backups_dir) if f.startswith('journal_backup_')],
            reverse=True
        )
        if files:
            filepath = os.path.join(backups_dir, files[0])
            size_mb = os.path.getsize(filepath) / 1024 / 1024
            send_document(chat_id, filepath,
                f'💾 Бэкап БД\n📅 {datetime.now().strftime("%d.%m.%Y %H:%M")}\n📦 {size_mb:.1f} МБ')
        else:
            send_message(chat_id, '❌ Бэкап создан, но файл не найден')
    except Exception as e:
        send_message(chat_id, f'❌ Ошибка: {e}')


def handle_stats(chat_id):
    """Получить статистику из приложения."""
    try:
        import sqlite3
        db_path = os.path.join(BASE_DIR, 'instance', 'journal.db')
        conn = sqlite3.connect(db_path)
        c = conn.cursor()

        total = c.execute('SELECT COUNT(*) FROM article').fetchone()[0]
        paid = c.execute('SELECT COUNT(*) FROM article WHERE payment_received = 1').fetchone()[0]
        reviewed = c.execute('SELECT COUNT(*) FROM article WHERE has_review = 1').fetchone()[0]
        edited = c.execute('SELECT COUNT(*) FROM article WHERE edited = 1').fetchone()[0]
        journals = c.execute('SELECT COUNT(*) FROM journal').fetchone()[0]
        issues = c.execute('SELECT COUNT(*) FROM issue').fetchone()[0]
        conn.close()

        send_message(chat_id,
            f'📈 <b>Статистика Radiotec Journal App</b>\n\n'
            f'📚 Журналов: <b>{journals}</b>\n'
            f'📖 Выпусков: <b>{issues}</b>\n'
            f'📄 Статей: <b>{total}</b>\n\n'
            f'💰 Оплачено: <b>{paid}</b> / {total}\n'
            f'📝 С рецензией: <b>{reviewed}</b> / {total}\n'
            f'✏️ Отредактировано: <b>{edited}</b> / {total}'
        )
    except Exception as e:
        send_message(chat_id, f'❌ Ошибка: {e}')


def is_authorized(chat_id):
    """Проверяет, авторизован ли пользователь."""
    return str(chat_id) in _authorized_users


def process_update(update):
    """Обработать одно обновление от Telegram."""
    msg = update.get('message')
    if not msg:
        return

    chat_id = msg['chat']['id']
    text = msg.get('text', '').strip()

    # /start доступен всем
    if text == '/start':
        if is_authorized(chat_id):
            handle_start(chat_id)
        else:
            send_message(chat_id,
                '🔒 <b>Radiotec Journal App</b>\n\n'
                'Для доступа введите пароль:'
            )
        return

    # Если не авторизован — проверяем пароль
    if not is_authorized(chat_id):
        if text == BOT_PASSWORD:
            _authorized_users[str(chat_id)] = True
            user_name = msg.get('from', {}).get('first_name', 'Пользователь')
            # Уведомляем владельца
            send_message(TELEGRAM_CHAT_ID,
                f'🔓 Новый пользователь получил доступ к боту:\n'
                f'👤 {user_name} (ID: {chat_id})')
            send_message(chat_id,
                '✅ Доступ разрешён! Добро пожаловать.',
                reply_markup=get_main_keyboard()
            )
        else:
            send_message(chat_id, '❌ Неверный пароль. Попробуйте ещё раз.')
        return

    # Авторизованные команды
    if text in ('/export', '📊 Excel-выгрузка'):
        handle_export(chat_id)
    elif text in ('/backup', '💾 Бэкап БД'):
        handle_backup(chat_id)
    elif text in ('/stats', '📈 Статистика'):
        handle_stats(chat_id)
    else:
        send_message(chat_id,
            'Используйте кнопки или команды:\n'
            '/export — Excel-выгрузка\n'
            '/backup — Бэкап БД\n'
            '/stats — Статистика',
            reply_markup=get_main_keyboard()
        )


def main():
    """Основной цикл polling."""
    print(f'[{datetime.now()}] Telegram bot started (polling)')
    offset = None

    while True:
        try:
            params = {'timeout': 30}
            if offset:
                params['offset'] = offset
            resp = requests.get(f'{API}/getUpdates', params=params, timeout=35)
            data = resp.json()

            if data.get('ok') and data.get('result'):
                for update in data['result']:
                    offset = update['update_id'] + 1
                    try:
                        process_update(update)
                    except Exception as e:
                        print(f'Error processing update: {e}')

        except requests.exceptions.Timeout:
            continue
        except Exception as e:
            print(f'Polling error: {e}')
            time.sleep(5)


if __name__ == '__main__':
    main()
