import csv
import io
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from functools import wraps
from io import BytesIO

from flask import (
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    flash,
    current_app,
    send_file,
)
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func, and_
from sqlalchemy.orm import joinedload
from werkzeug.utils import secure_filename

from models import db, User, Journal, Issue, Article, ArticleAuthor, ArticleImage, ActivityLog, ArticleHistory, ArticleComment


ALLOWED_EXTENSIONS = {
    # Документы
    'txt', 'pdf', 'doc', 'docx', 'odt', 'rtf', 'xls', 'xlsx', 'csv', 'ppt', 'pptx',
    # Изображения
    'jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'svg', 'tiff', 'tif', 'ico',
    # Архивы
    'zip', 'rar', '7z', 'tar', 'gz',
    # Аудио/видео
    'mp3', 'wav', 'mp4', 'avi', 'mov', 'mkv',
    # Прочее
    'html', 'htm', 'xml', 'json', 'yaml', 'yml', 'md', 'tex', 'djvu', 'epub', 'fb2',
}

MAX_UPLOAD_SIZE_MB = 100
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.is_admin:
            flash('У вас нет доступа к админ-панели', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated


def log_activity(action, entity_type, entity_id=None, entity_title=None, details=None):
    """Записывает действие в журнал активности."""
    entry = ActivityLog(
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_title=entity_title,
        details=details,
        created_at=datetime.now(timezone.utc)
    )
    db.session.add(entry)


def log_article_history(article_id, action, changes=None, user_name=None):
    """Записывает изменение статьи в историю."""
    if user_name is None:
        user_name = current_user.display_name if current_user and current_user.is_authenticated else 'Система'
    entry = ArticleHistory(
        article_id=article_id,
        user_name=user_name,
        action=action,
        changes=json.dumps(changes, ensure_ascii=False) if changes else None,
        created_at=datetime.now(timezone.utc)
    )
    db.session.add(entry)


def _read_sample_and_rewind(file, sample_size=8192):
    """Читает первые байты файла и пытается вернуть указатель в начало."""
    stream = getattr(file, "stream", None)
    if stream is None:
        return None
    try:
        pos = stream.tell()
        sample = stream.read(sample_size)
        # Надо обязательно перемотать обратно, иначе `file.save()` сохранит урезанный файл.
        stream.seek(pos)
        return sample
    except Exception:
        return None


def _detect_mime_from_signature(sample: bytes):
    """Пытается определить MIME по magic-signature (без внешних зависимостей)."""
    if not sample:
        return None

    # PDF
    if sample.startswith(b"%PDF-"):
        return "application/pdf"

    # PNG
    if sample.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"

    # JPEG
    if sample.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"

    # GIF
    if sample.startswith(b"GIF87a") or sample.startswith(b"GIF89a"):
        return "image/gif"

    # WEBP: RIFF....WEBP
    if sample.startswith(b"RIFF") and b"WEBP" in sample[8:32]:
        return "image/webp"

    # BMP
    if sample.startswith(b"BM"):
        return "image/bmp"

    # TIFF
    if sample.startswith(b"II*\x00") or sample.startswith(b"MM\x00*"):
        return "image/tiff"

    # SVG: ищем именно `<svg` (а не любой XML, начинающийся с `<?xml`)
    stripped = sample.lstrip()
    sample_lower = sample[:2048].lower()
    if stripped.startswith(b"<svg") or (stripped.startswith(b"<?xml") and b"<svg" in sample_lower):
        return "image/svg+xml"

    # RTF
    if sample.startswith(b"{\\rtf"):
        return "application/rtf"

    # OLE DOC (.doc)
    if sample.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        # У многих форматов MS Office (doc/xls/ppt) используется общий OLE-заголовок.
        return "application/x-ole-storage"

    # ZIP-based formats (.docx/.xlsx/.pptx/.odt/.epub и т.п.)
    if sample.startswith(b"PK\x03\x04") or sample.startswith(b"PK\x05\x06") or sample.startswith(b"PK\x07\x08"):
        return "application/zip"

    # 7z
    if sample.startswith(b"7z\xBC\xAF\x27\x1C"):
        return "application/x-7z-compressed"

    # RAR
    if sample.startswith(b"Rar!\x1a\x07\x00") or sample.startswith(b"Rar!\x1a\x07") or sample.startswith(b"Rar!\x1a"):
        return "application/x-rar-compressed"

    # GZIP
    if sample.startswith(b"\x1f\x8b"):
        return "application/gzip"

    # TAR: в заголовке по смещению 257 лежит 'ustar'
    if len(sample) > 262 and sample[257:262] == b"ustar":
        return "application/x-tar"

    # Текстовые файлы (очень грубая эвристика)
    if b"\x00" not in sample:
        try:
            sample.decode("utf-8")
            return "text/plain"
        except UnicodeDecodeError:
            try:
                sample.decode("cp1251")
                return "text/plain"
            except UnicodeDecodeError:
                return None

    return None


def _expected_mime_for_extension(ext: str):
    """Ожидаемый MIME для расширения (для тех, где можно проверить по сигнатуре)."""
    ext = ext.lower().lstrip(".")
    mapping = {
        "pdf": "application/pdf",
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "gif": "image/gif",
        "webp": "image/webp",
        "bmp": "image/bmp",
        "tif": "image/tiff",
        "tiff": "image/tiff",
        "svg": "image/svg+xml",
        "rtf": "application/rtf",
        "doc": "application/x-ole-storage",
        # ZIP-based office formats
        "docx": "application/zip",
        # старые xls/ppt тоже OLE-сущность
        "xls": "application/x-ole-storage",
        "xlsx": "application/zip",
        "ppt": "application/x-ole-storage",
        "pptx": "application/zip",
        "odt": "application/zip",
        "ods": "application/zip",
        "odp": "application/zip",
        "epub": "application/zip",
        "tex": "text/plain",
        "fb2": "text/plain",
        "csv": "text/plain",
        "txt": "text/plain",
        "json": "text/plain",
        "yaml": "text/plain",
        "yml": "text/plain",
        "md": "text/plain",
        "xml": "text/plain",
        "html": "text/plain",
        "htm": "text/plain",
        "zip": "application/zip",
        "rar": "application/x-rar-compressed",
        "7z": "application/x-7z-compressed",
        "gz": "application/gzip",
        "tar": "application/x-tar",
    }
    return mapping.get(ext)


def _save_uploaded_file(file, upload_folder, field_label=None):
    """Сохраняет загруженный файл и возвращает имя файла, или (None, error)."""
    if not file or not getattr(file, "filename", None):
        return None, None

    original_name = file.filename
    if not allowed_file(original_name):
        return None, f"Файл \"{original_name}\" имеет неподдерживаемое расширение."

    # Проверка размера (100MB)
    size = getattr(file, "content_length", None)
    if size is None:
        # Иногда `content_length` может быть None — тогда попробуем узнать размер через stream.
        try:
            stream = file.stream
            pos = stream.tell()
            stream.seek(0, 2)
            size = stream.tell()
            stream.seek(pos)
        except Exception:
            size = None
    if size is not None and size > MAX_UPLOAD_SIZE_BYTES:
        label = f"{field_label}: " if field_label else ""
        return None, f"{label}Файл слишком большой (максимум {MAX_UPLOAD_SIZE_MB}MB): \"{original_name}\""

    ext = original_name.rsplit(".", 1)[1].lower() if "." in original_name else ""
    expected_mime = _expected_mime_for_extension(ext)

    if expected_mime:
        sample = _read_sample_and_rewind(file)
        detected_mime = _detect_mime_from_signature(sample)
        if detected_mime != expected_mime:
            label = f"{field_label}: " if field_label else ""
            return None, (
                f"{label}Недопустимый MIME-тип файла \"{original_name}\": "
                f"обнаружено {detected_mime or 'unknown'}, ожидалось {expected_mime}."
            )

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
    filename = secure_filename(timestamp + original_name)
    file.save(os.path.join(upload_folder, filename))
    return filename, None


def _process_file_uploads(article, upload_folder):
    """Обрабатывает загрузку файлов (рукопись, рецензия, титульная, акт экспертизы) для статьи."""
    errors = []
    for field_name in ('manuscript_file', 'review_file', 'title_pdf', 'expertise_act_file'):
        file = request.files.get(field_name)
        field_labels = {
            "manuscript_file": "Текст статьи",
            "review_file": "Рецензия",
            "title_pdf": "Титульная",
            "expertise_act_file": "Акт экспертизы",
        }
        saved, err = _save_uploaded_file(file, upload_folder, field_label=field_labels.get(field_name, field_name))
        if err:
            errors.append(err)
        if saved:
            setattr(article, field_name, saved)
            # Автоматически включаем статус акта при загрузке файла
            if field_name == 'expertise_act_file':
                article.has_expertise_act = True
    return errors


def _process_article_images(article, upload_folder, start_order=0):
    """Загружает файлы раздела 'Разное' для статьи. Возвращает количество добавленных."""
    images = request.files.getlist('article_images')
    added = 0
    errors = []
    for i, img in enumerate(images):
        saved, err = _save_uploaded_file(img, upload_folder, field_label="Изображение")
        if err:
            errors.append(err)
        if saved:
            db.session.add(ArticleImage(
                article_id=article.id,
                filename=saved,
                order=start_order + i
            ))
            added += 1
    return added, errors


def _process_authors(article):
    """Парсит авторов из формы и привязывает к статье. Возвращает строку с ФИО."""
    author_names = request.form.getlist('author_name[]')
    author_emails = request.form.getlist('author_email[]')
    author_orgs = request.form.getlist('author_organization[]')
    author_degrees = request.form.getlist('author_degree[]')
    author_positions = request.form.getlist('author_position[]')
    author_phones = request.form.getlist('author_phone[]')

    names_list = []
    for i, name in enumerate(author_names):
        if not name:
            continue
        names_list.append(name)
        db.session.add(ArticleAuthor(
            article_id=article.id,
            full_name=name,
            email=author_emails[i] if i < len(author_emails) else '',
            organization=author_orgs[i] if i < len(author_orgs) else '',
            degree=author_degrees[i] if i < len(author_degrees) else '',
            position=author_positions[i] if i < len(author_positions) else '',
            phone=author_phones[i] if i < len(author_phones) else '',
            order=i
        ))

    return ", ".join(names_list) if names_list else None


def _filtered_articles_query(journal_id=None, status_filter='', search=''):
    """Общий запрос статей с фильтрами. Только неудалённые статьи и выпуски."""
    query = (
        Article.visible()
        .join(Issue).join(Journal)
        .filter(Issue.deleted_at.is_(None))
        .options(
            joinedload(Article.issue).joinedload(Issue.journal),
            joinedload(Article.article_authors)
        )
    )
    if journal_id:
        query = query.filter(Issue.journal_id == journal_id)
    if status_filter == 'unpaid':
        query = query.filter(Article.payment_received == False)
    elif status_filter == 'no_review':
        query = query.filter(Article.has_review == False)
    elif status_filter == 'not_edited':
        query = query.filter(Article.edited == False)
    elif status_filter == 'no_expertise':
        query = query.filter(Article.has_expertise_act == False)
    if search:
        like_pattern = f'%{search}%'
        query = query.filter(
            db.or_(
                Article.title.ilike(like_pattern),
                Article.authors.ilike(like_pattern)
            )
        )
    return query.order_by(Article.id.desc())


def _build_filter_description(journal_id=None, status_filter='', search=''):
    """Строит текстовое описание применённых фильтров для экспорта."""
    parts = []
    if journal_id:
        journal = Journal.query.get(journal_id)
        if journal:
            parts.append(f'Журнал: {journal.name}')
    status_names = {'unpaid': 'Без оплаты', 'no_review': 'Без рецензии', 'not_edited': 'Не отредактировано', 'no_expertise': 'Без акта эксп.'}
    if status_filter in status_names:
        parts.append(f'Статус: {status_names[status_filter]}')
    if search:
        parts.append(f'Поиск: «{search}»')
    return ', '.join(parts) if parts else 'Все статьи'


def register_admin_routes(app):
    @app.route('/journal/<int:journal_id>/issues/reorder', methods=['POST'])
    @login_required
    def reorder_issues(journal_id):
        data = request.json['order']
        for new_pos, issue_id in enumerate(data):
            issue = Issue.visible().filter_by(id=issue_id, journal_id=journal_id).first()
            if issue:
                issue.position = new_pos
        db.session.commit()
        return jsonify({'status': 'ok'})

    @app.route("/journal/<int:journal_id>/issue/new", methods=['POST'])
    @login_required
    def new_issue(journal_id):
        number = request.form['number']
        year = request.form['year']
        issue = Issue(number=int(number), year=int(year), journal_id=journal_id)
        db.session.add(issue)
        db.session.flush()
        journal = Journal.query.get(journal_id)
        log_activity('created', 'issue', issue.id, f'№{number}/{year}', f'Журнал: {journal.name}' if journal else None)
        db.session.commit()
        flash(f'Выпуск №{number}/{year} создан', 'success')
        return redirect(f'/journal/{journal_id}')

    @app.route("/journal/<int:journal_id>/article/<int:article_id>/assign", methods=['POST'])
    @login_required
    @admin_required
    def assign_article_to_issue(journal_id, article_id):
        """Перенос статьи от автора в выбранный выпуск (или в нулевой). После этого статья как добавленная вручную."""
        issue_id = request.form.get('issue_id', type=int)
        if not issue_id:
            flash('Укажите выпуск', 'error')
            return redirect(url_for('journal_page', journal_id=journal_id))
        article = Article.visible().filter_by(id=article_id).first_or_404()
        issue = Issue.query.filter_by(id=issue_id, journal_id=journal_id).first_or_404()
        article.issue_id = issue_id
        # submitted_by_user_id не сбрасываем — статья остаётся видна автору в кабинете
        db.session.commit()
        log_activity('updated', 'article', article.id, article.title, f'Перенесена в выпуск №{issue.number}/{issue.year}')
        flash('Статья перенесена в выпуск №{} / {}'.format(issue.number, issue.year), 'success')
        return redirect(url_for('journal_page', journal_id=journal_id))

    @app.route("/issue/<int:issue_id>/add-article", methods=['POST'])
    @login_required
    def add_article(issue_id):
        title = request.form.get('title')
        notes = request.form.get('notes', '')

        # Формирование даты поступления
        day = request.form.get('submission_day', '')
        month = request.form.get('submission_month', '')
        year = request.form.get('submission_year', '')
        submission_date = f"{day}.{month}.{year}" if day and month and year else None

        article = Article(
            title=title,
            submission_date=submission_date,
            notes=notes,
            issue_id=issue_id
        )

        upload_folder = current_app.config['UPLOAD_FOLDER']
        upload_errors = _process_file_uploads(article, upload_folder)
        for msg in upload_errors:
            flash(msg, 'error')

        # Если рукопись была загружена через PDF-парсер
        parsed_manuscript = request.form.get('parsed_manuscript_filename')
        if parsed_manuscript and not article.manuscript_file:
            # Проверяем что файл действительно существует
            if os.path.exists(os.path.join(upload_folder, parsed_manuscript)):
                article.manuscript_file = parsed_manuscript

        db.session.add(article)
        db.session.flush()

        _, image_errors = _process_article_images(article, upload_folder)
        for msg in image_errors:
            flash(msg, 'error')
        article.authors = _process_authors(article)

        log_activity('created', 'article', article.id, article.title)
        log_article_history(article.id, 'created', [{'field': 'Статья создана', 'old': '', 'new': article.title}])
        db.session.commit()
        flash('Статья добавлена', 'success')
        return redirect(f'/issue/{issue_id}')

    @app.route("/article/<int:article_id>/toggle/<field>", methods=['POST'])
    @login_required
    def toggle_article(article_id, field):
        article = Article.visible().filter_by(id=article_id).first_or_404()
        issue_id = article.issue_id
        new_value = False

        if field == 'payment':
            article.payment_received = not article.payment_received
            new_value = article.payment_received
        elif field == 'edited':
            article.edited = not article.edited
            new_value = article.edited
        elif field == 'review':
            article.has_review = not article.has_review
            new_value = article.has_review
        elif field == 'expertise':
            article.has_expertise_act = not article.has_expertise_act
            new_value = article.has_expertise_act

        field_labels = {'payment': 'Оплата', 'edited': 'Редакт.', 'review': 'Рецензия', 'expertise': 'Акт эксп.'}
        status_text = 'вкл' if new_value else 'выкл'
        log_activity('toggled', 'article', article.id, article.title, f'{field_labels.get(field, field)}: {status_text}')
        log_article_history(article.id, 'status', [{'field': field_labels.get(field, field), 'old': 'выкл' if new_value else 'вкл', 'new': status_text}])
        db.session.commit()
        
        # Если AJAX-запрос, возвращаем JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'field': field, 'value': new_value})
        
        return redirect(f'/issue/{issue_id}')

    @app.route('/article/<int:article_id>/update-notes', methods=['POST'])
    @login_required
    def update_notes(article_id):
        data = request.json
        article = Article.visible().filter_by(id=article_id).first()
        if article:
            article.notes = data.get('notes', '')
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'success': False}), 404

    @app.route("/article/<int:article_id>/delete", methods=['POST'])
    @login_required
    def delete_article(article_id):
        article = Article.visible().filter_by(id=article_id).first_or_404()
        issue_id = article.issue_id
        log_activity('deleted', 'article', article.id, article.title)
        article.deleted_at = datetime.now(timezone.utc)
        db.session.commit()
        flash('Статья перемещена в корзину', 'success')
        return redirect(f'/issue/{issue_id}')

    @app.route("/issue/<int:issue_id>/delete", methods=['POST'])
    @login_required
    def delete_issue(issue_id):
        issue = Issue.visible().options(joinedload(Issue.journal)).filter_by(id=issue_id).first_or_404()
        journal_id = issue.journal_id
        log_activity('deleted', 'issue', issue.id, f'№{issue.number}/{issue.year}', f'Журнал: {issue.journal.name}')
        now = datetime.now(timezone.utc)
        issue.deleted_at = now
        for article in issue.articles:
            article.deleted_at = now
        db.session.commit()
        flash(f'Выпуск №{issue.number}/{issue.year} перемещён в корзину', 'success')
        return redirect(f'/journal/{journal_id}')

    @app.route("/article/<int:article_id>/edit", methods=['GET', 'POST'])
    @login_required
    def edit_article(article_id):
        article = Article.visible().options(
            joinedload(Article.article_authors),
            joinedload(Article.images)
        ).filter_by(id=article_id).first_or_404()
        issue = Issue.query.options(joinedload(Issue.journal)).get(article.issue_id)
        journal = issue.journal if issue else None
        
        if request.method == 'POST':
            # Отслеживаем изменения
            changes = []
            new_title = request.form.get('title', article.title)
            new_notes = request.form.get('notes', '')
            if new_title != article.title:
                changes.append({'field': 'Название', 'old': article.title or '', 'new': new_title})
            if new_notes != (article.notes or ''):
                changes.append({'field': 'Заметки', 'old': (article.notes or '')[:100], 'new': new_notes[:100]})

            # Основные поля
            article.title = new_title
            article.notes = new_notes
            
            # Дата поступления
            day = request.form.get('submission_day', '')
            month = request.form.get('submission_month', '')
            year = request.form.get('submission_year', '')
            if day and month and year:
                new_date = f"{day}.{month}.{year}"
                if new_date != (article.submission_date or ''):
                    changes.append({'field': 'Дата поступления', 'old': article.submission_date or '', 'new': new_date})
                article.submission_date = new_date
            
            # Статусы — только при обычной отправке формы (не AJAX из slide-panel)
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if not is_ajax:
                new_statuses = {
                    'payment_received': ('payment_received' in request.form, 'Оплата'),
                    'has_review': ('has_review' in request.form, 'Рецензия'),
                    'edited': ('edited' in request.form, 'Редактирование'),
                    'has_expertise_act': ('has_expertise_act' in request.form, 'Акт эксп.'),
                }
                for field, (new_val, label) in new_statuses.items():
                    old_val = getattr(article, field)
                    if old_val != new_val:
                        changes.append({'field': label, 'old': 'вкл' if old_val else 'выкл', 'new': 'вкл' if new_val else 'выкл'})
                    setattr(article, field, new_val)
            
            # Загрузка файлов
            upload_folder = current_app.config['UPLOAD_FOLDER']
            upload_errors = _process_file_uploads(article, upload_folder)
            if upload_errors:
                error_message = "; ".join(upload_errors)
                if is_ajax:
                    return jsonify({'success': False, 'error': error_message}), 400
                for msg in upload_errors:
                    flash(msg, 'error')
                return redirect(url_for('edit_article', article_id=article.id))
            
            # Удаление файлов
            for field_name, form_key in [('manuscript_file', 'delete_manuscript'), ('review_file', 'delete_review'), ('title_pdf', 'delete_title_pdf'), ('expertise_act_file', 'delete_expertise_act')]:
                if request.form.get(form_key):
                    setattr(article, field_name, None)
                    # Сбрасываем статус акта при удалении файла
                    if field_name == 'expertise_act_file':
                        article.has_expertise_act = False
            
            # Удаление выбранных изображений
            delete_image_ids = request.form.getlist('delete_image[]')
            for img_id in delete_image_ids:
                img = ArticleImage.query.get(int(img_id))
                if img and img.article_id == article.id:
                    db.session.delete(img)
            
            # Загрузка новых изображений
            current_max_order = max([img.order for img in article.images], default=-1)
            _, image_errors = _process_article_images(article, upload_folder, start_order=current_max_order + 1)
            if image_errors:
                upload_errors.extend(image_errors)

            if upload_errors:
                error_message = "; ".join(upload_errors)
                if is_ajax:
                    return jsonify({'success': False, 'error': error_message}), 400
                for msg in upload_errors:
                    flash(msg, 'error')
                return redirect(url_for('edit_article', article_id=article.id))
            
            # Обновление авторов — удаляем старых и добавляем новых
            old_authors = article.authors or ''
            ArticleAuthor.query.filter_by(article_id=article.id).delete()
            new_authors_str = _process_authors(article)

            if old_authors != (new_authors_str or ''):
                changes.append({'field': 'Авторы', 'old': old_authors[:100], 'new': (new_authors_str or '')[:100]})
            article.authors = new_authors_str

            log_activity('updated', 'article', article.id, article.title)
            if changes:
                log_article_history(article.id, 'updated', changes)
            db.session.commit()

            # AJAX-ответ для slide-panel
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Собираем обновлённые данные для карточки
                authors_list = []
                for a in sorted(article.article_authors, key=lambda x: x.order):
                    authors_list.append({'name': a.full_name, 'degree': a.degree or ''})
                return jsonify({
                    'success': True,
                    'message': 'Статья обновлена',
                    'article': {
                        'id': article.id,
                        'title': article.title,
                        'authors': authors_list,
                        'authors_str': article.authors or '',
                        'submission_date': article.submission_date or '',
                        'payment': article.payment_received,
                        'review': article.has_review,
                        'edited': article.edited,
                        'expertise': article.has_expertise_act,
                        'expertise_act_file': article.expertise_act_file or '',
                        'manuscript_file': article.manuscript_file or '',
                        'review_file': article.review_file or '',
                        'title_pdf': article.title_pdf or '',
                        'notes': article.notes or '',
                    }
                })

            flash('Статья обновлена', 'success')
            return redirect(f'/issue/{article.issue_id}')
        
        return render_template('edit_article.html', article=article, issue=issue, journal=journal)

    @app.route('/article/<int:article_id>/move', methods=['POST'])
    @login_required
    def move_article(article_id):
        data = request.get_json()
        new_issue_id = data.get('issue_id')

        article = Article.visible().filter_by(id=article_id).first()
        if not article:
            return jsonify({'error': 'Статья не найдена'}), 404

        issue = Issue.visible().filter_by(id=new_issue_id).first()
        if not issue:
            return jsonify({'error': 'Выпуск не найден'}), 400

        old_issue = article.issue
        old_label = f'{old_issue.journal.name} №{old_issue.number}/{old_issue.year}'
        new_label = f'{issue.journal.name} №{issue.number}/{issue.year}'
        article.issue_id = new_issue_id
        log_article_history(article.id, 'moved', [{'field': 'Выпуск', 'old': old_label, 'new': new_label}])
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Статья перемещена из "{old_label}" в "{new_label}"'
        })

    @app.route('/issue/<int:issue_id>/export-excel')
    @login_required
    def export_issue_excel(issue_id):
        """Экспортирует все статьи выпуска в Excel"""
        issue = Issue.visible().filter_by(id=issue_id).first()
        if not issue:
            return "Выпуск не найден", 404

        # Создаём Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Выпуск {issue.number}"

        # Стили
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Заголовки
        headers = ["№", "Название статьи", "Авторы", "Дата поступления", "Оплачено", "Рецензия", "Редактировано"]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные статей (для выпуска «без номера» — без авторских подач, они только в «От авторов»)
        articles_query = Article.visible().filter_by(issue_id=issue_id).options(joinedload(Article.article_authors))
        if issue.number == 0 and issue.year == 0:
            articles_query = articles_query.filter(Article.submitted_by_user_id.is_(None))
        articles_list = articles_query.all()
        for idx, article in enumerate(articles_list, 1):
            authors = ", ".join([f"{a.full_name}" for a in article.article_authors]) if article.article_authors else article.authors or "-"

            ws.append([
                idx,
                article.title or "-",
                authors,
                article.submission_date or "-",
                "✓" if article.payment_received else "✗",
                "✓" if article.has_review else "✗",
                "✓" if article.edited else "✗"
            ])

        # Ширина колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15

        # Отправляем файл
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Выпуск_{issue.number}_{issue.year}.xlsx"
        )

    @app.route('/issue/<int:issue_id>/export-csv')
    @login_required
    def export_issue_csv(issue_id):
        """Экспортирует все статьи выпуска в CSV"""
        issue = Issue.visible().filter_by(id=issue_id).first()
        if not issue:
            return "Выпуск не найден", 404

        # Создаём CSV в памяти
        output = io.StringIO()
        writer = csv.writer(output)

        # Заголовки
        headers = ["№", "Название статьи", "Авторы", "Дата поступления", "Оплачено", "Рецензия", "Редактировано"]
        writer.writerow(headers)

        # Данные статей (для выпуска «без номера» — без авторских подач)
        articles_query = Article.visible().filter_by(issue_id=issue_id).options(joinedload(Article.article_authors))
        if issue.number == 0 and issue.year == 0:
            articles_query = articles_query.filter(Article.submitted_by_user_id.is_(None))
        articles_list = articles_query.all()
        for idx, article in enumerate(articles_list, 1):
            authors = ", ".join([f"{a.full_name}" for a in article.article_authors]) if article.article_authors else article.authors or "-"

            writer.writerow([
                idx,
                article.title or "-",
                authors,
                article.submission_date or "-",
                "✓" if article.payment_received else "✗",
                "✓" if article.has_review else "✗",
                "✓" if article.edited else "✗"
            ])

        # Конвертируем в BytesIO для отправки
        output.seek(0)
        bytes_output = BytesIO(output.getvalue().encode('utf-8-sig'))

        return send_file(
            bytes_output,
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name=f"Выпуск_{issue.number}_{issue.year}.csv"
        )

    @app.route('/admin')
    @admin_required
    def admin_home():
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/logout')
    @login_required
    def admin_logout():
        return redirect(url_for('logout'))

    @app.route('/admin/dashboard')
    @admin_required
    def admin_dashboard():
        try:
            total_articles = Article.visible().count()
            unpaid = Article.visible().filter_by(payment_received=False).count()
            no_review = Article.visible().filter_by(has_review=False).count()
            not_edited = Article.visible().filter_by(edited=False).count()
            no_expertise = Article.visible().filter_by(has_expertise_act=False).count()
            
            stats = {
                'journals': Journal.query.count(),
                'issues': Issue.visible().count(),
                'articles': total_articles,
                'unpaid': unpaid,
                'no_review': no_review,
                'not_edited': not_edited,
                'no_expertise': no_expertise,
            }
            recent_articles = (
                Article.visible()
                .options(joinedload(Article.issue).joinedload(Issue.journal))
                .order_by(Article.id.desc())
                .limit(10)
                .all()
            )
            recent_activity = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(15).all()
            
            # Данные для графика: статьи по журналам (только неудалённые)
            chart_query = (
                db.session.query(Journal.name, func.count(Article.id))
                .join(Issue, Issue.journal_id == Journal.id)
                .join(Article, Article.issue_id == Issue.id)
                .filter(Issue.deleted_at.is_(None), Article.deleted_at.is_(None))
                .group_by(Journal.id, Journal.name)
                .having(func.count(Article.id) > 0)
                .order_by(func.count(Article.id).desc())
                .all()
            )
            chart_data = [{'name': name, 'count': count} for name, count in chart_query]
            max_articles = chart_data[0]['count'] if chart_data else 0
            
            # Данные для doughnut-графика: статусы статей
            status_data = {
                'paid': total_articles - unpaid,
                'unpaid': unpaid,
                'reviewed': total_articles - no_review,
                'no_review': no_review,
                'edited': total_articles - not_edited,
                'not_edited': not_edited,
                'with_expertise': total_articles - no_expertise,
                'no_expertise': no_expertise,
            }
            
            
        except Exception:
            stats = {'journals': 0, 'issues': 0, 'articles': 0, 'unpaid': 0, 'no_review': 0, 'not_edited': 0, 'no_expertise': 0}
            recent_articles = []
            recent_activity = []
            chart_data = []
            max_articles = 0
            status_data = {'paid': 0, 'unpaid': 0, 'reviewed': 0, 'no_review': 0, 'edited': 0, 'not_edited': 0, 'with_expertise': 0, 'no_expertise': 0}

        return render_template('admin/dashboard.html', 
                               stats=stats, 
                               recent_articles=recent_articles,
                               recent_activity=recent_activity,
                               chart_data=chart_data,
                               max_articles=max_articles,
                               status_data=status_data)

    @app.route('/admin/journals')
    @admin_required
    def admin_journals():
        """Страница управления журналами"""
        # Один запрос для всех журналов с подсчётами (вместо 2N запросов)
        journals_with_counts = (
            db.session.query(
                Journal,
                func.count(func.distinct(Issue.id)).label('issue_count'),
                func.count(Article.id).label('article_count')
            )
            .outerjoin(Issue, and_(Issue.journal_id == Journal.id, Issue.deleted_at.is_(None)))
            .outerjoin(Article, and_(Article.issue_id == Issue.id, Article.deleted_at.is_(None)))
            .group_by(Journal.id)
            .order_by(Journal.id)
            .all()
        )
        journals_data = []
        for j, issue_count, article_count in journals_with_counts:
            journals_data.append({
                'id': j.id,
                'is_hidden': j.is_hidden or False,
                'name': j.name,
                'issn': j.issn or '',
                'issues': issue_count,
                'articles': article_count,
            })
        return render_template('admin/journals.html', journals=journals_data)

    @app.route('/admin/journals/add', methods=['POST'])
    @admin_required
    def admin_journal_add():
        """Добавление нового журнала"""
        data = request.get_json()
        name = (data.get('name') or '').strip()
        issn = (data.get('issn') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': 'Название не может быть пустым'}), 400
        journal = Journal(name=name, issn=issn or None)
        db.session.add(journal)
        db.session.flush()
        log_activity('created', 'journal', journal.id, journal.name)
        db.session.commit()
        return jsonify({'success': True, 'id': journal.id, 'name': journal.name, 'issn': journal.issn or ''})

    @app.route('/admin/journals/<int:journal_id>/rename', methods=['POST'])
    @admin_required
    def admin_journal_rename(journal_id):
        """Переименование журнала"""
        data = request.get_json()
        name = (data.get('name') or '').strip()
        issn = data.get('issn')
        if not name:
            return jsonify({'success': False, 'message': 'Название не может быть пустым'}), 400
        journal = Journal.query.get_or_404(journal_id)
        old_name = journal.name
        journal.name = name
        if issn is not None:
            journal.issn = issn.strip() or None
        log_activity('updated', 'journal', journal.id, journal.name, f'Было: {old_name}')
        db.session.commit()
        return jsonify({'success': True, 'name': journal.name, 'issn': journal.issn or ''})

    @app.route('/admin/journals/<int:journal_id>/toggle-visibility', methods=['POST'])
    @admin_required
    def admin_journal_toggle_visibility(journal_id):
        """Скрыть/показать журнал на публичной странице"""
        journal = Journal.query.get_or_404(journal_id)
        journal.is_hidden = not journal.is_hidden
        log_activity('updated', 'journal', journal.id, journal.name,
                      'Скрыт' if journal.is_hidden else 'Показан')
        db.session.commit()
        return jsonify({'success': True, 'is_hidden': journal.is_hidden})

    @app.route('/admin/journals/<int:journal_id>/delete', methods=['POST'])
    @admin_required
    def admin_journal_delete(journal_id):
        """Удаление журнала"""
        journal = Journal.query.get_or_404(journal_id)
        # Проверяем наличие выпусков
        issue_count = Issue.visible().filter_by(journal_id=journal_id).count()
        if issue_count > 0:
            return jsonify({
                'success': False,
                'message': f'Невозможно удалить: журнал содержит {issue_count} выпуск(ов). Сначала удалите все выпуски.'
            }), 400
        log_activity('deleted', 'journal', journal.id, journal.name)
        db.session.delete(journal)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/admin/articles')
    @admin_required
    def admin_articles():
        """Страница управления всеми статьями — начальная загрузка без статей, всё через AJAX"""
        journal_id = request.args.get('journal_id', type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '').strip()
        
        journals = Journal.query.all()
        total_articles = Article.visible().count()
        
        return render_template('admin/articles.html', 
                               journals=journals,
                               total_articles=total_articles,
                               current_journal=journal_id,
                               current_status=status_filter,
                               search=search)

    @app.route('/admin/articles/api')
    @admin_required
    def admin_articles_api():
        """JSON API для AJAX-фильтрации статей с пагинацией"""
        journal_id = request.args.get('journal_id', type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 50, type=int), 200)

        query = _filtered_articles_query(journal_id, status_filter, search)
        total = query.count()
        articles = query.offset((page - 1) * per_page).limit(per_page).all()

        result = []
        for a in articles:
            authors_str = ''
            if a.article_authors:
                authors_str = ', '.join(au.full_name for au in sorted(a.article_authors, key=lambda x: x.order))
            else:
                authors_str = a.authors or '-'

            journal_name = ''
            issue_label = ''
            issue_id_val = None
            if a.issue:
                issue_id_val = a.issue_id
                if a.issue.journal:
                    journal_name = a.issue.journal.name
                issue_label = f'\u2116{a.issue.number}/{a.issue.year}'

            result.append({
                'id': a.id,
                'title': a.title or '',
                'authors': authors_str,
                'journal_name': journal_name,
                'issue_label': issue_label,
                'issue_id': issue_id_val,
                'payment': a.payment_received,
                'review': a.has_review,
                'edited': a.edited,
                'expertise': a.has_expertise_act,
                'expertise_act_file': a.expertise_act_file or '',
            })

        total_pages = (total + per_page - 1) // per_page
        return jsonify({
            'articles': result,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
        })

    @app.route('/admin/articles/bulk-delete', methods=['POST'])
    @admin_required
    def bulk_delete_articles():
        """Массовое удаление статей"""
        data = request.get_json()
        article_ids = data.get('ids', [])
        
        if not article_ids:
            return jsonify({'success': False, 'message': 'Не выбрано ни одной статьи'}), 400
        
        now = datetime.now(timezone.utc)
        deleted_count = 0
        for aid in article_ids:
            article = Article.query.get(aid)  # любой, в т.ч. из корзины — проверяем deleted_at
            if article and not article.deleted_at:
                article.deleted_at = now
                deleted_count += 1

        db.session.commit()
        return jsonify({'success': True, 'message': f'В корзину перемещено статей: {deleted_count}'})

    @app.route('/admin/articles/bulk-toggle', methods=['POST'])
    @admin_required
    def bulk_toggle_articles():
        """Массовое изменение статуса статей"""
        data = request.get_json()
        article_ids = data.get('ids', [])
        field = data.get('field', '')  # payment, review, edited
        value = data.get('value', True)
        
        if not article_ids:
            return jsonify({'success': False, 'message': 'Не выбрано ни одной статьи'}), 400
        
        updated_count = 0
        for article_id in article_ids:
            article = Article.visible().filter_by(id=article_id).first()
            if article:
                if field == 'payment':
                    article.payment_received = value
                elif field == 'review':
                    article.has_review = value
                elif field == 'edited':
                    article.edited = value
                elif field == 'expertise':
                    article.has_expertise_act = value
                updated_count += 1
        
        db.session.commit()
        
        field_names = {'payment': 'оплата', 'review': 'рецензия', 'edited': 'редактирование', 'expertise': 'акт экспертизы'}
        status = 'установлен' if value else 'снят'
        return jsonify({'success': True, 'message': f'{field_names.get(field, field)}: {status} для {updated_count} статей'})

    @app.route('/admin/articles/bulk-export')
    @admin_required
    def bulk_export_articles():
        """Экспорт статей в Excel (с фильтрами или по ID)."""
        from excel_export import generate_articles_excel

        article_ids = request.args.get('ids', '')
        journal_id = request.args.get('journal_id', type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '').strip()

        if article_ids:
            ids_list = [int(x) for x in article_ids.split(',') if x.isdigit()]
            articles = (
                Article.visible()
                .filter(Article.id.in_(ids_list))
                .options(
                    joinedload(Article.issue).joinedload(Issue.journal),
                    joinedload(Article.article_authors)
                )
                .all()
            )
            filter_desc = f'Выбранные статьи ({len(ids_list)} шт.)'
        else:
            articles = _filtered_articles_query(journal_id, status_filter, search).all()
            filter_desc = _build_filter_description(journal_id, status_filter, search)

        # Преобразуем ORM-объекты в формат для общего модуля
        rows = []
        for article in articles:
            authors = ", ".join([a.full_name for a in article.article_authors]) if article.article_authors else (article.authors or "-")
            rows.append({
                'title': article.title,
                'authors': authors,
                'journal_name': article.issue.journal.name if article.issue and article.issue.journal else '-',
                'issue_info': f"№{article.issue.number}/{article.issue.year}" if article.issue else '-',
                'submission_date': article.submission_date,
                'payment': article.payment_received,
                'review': article.has_review,
                'edited': article.edited,
                'expertise': article.has_expertise_act,
            })

        output, total = generate_articles_excel(rows, filter_desc)

        filename = f"articles_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)

    @app.route('/admin/articles/bulk-export-csv')
    @admin_required
    def bulk_export_articles_csv():
        """Экспорт статей в CSV (с фильтрами или по ID)."""
        article_ids = request.args.get('ids', '')
        journal_id = request.args.get('journal_id', type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '').strip()

        if article_ids:
            ids_list = [int(x) for x in article_ids.split(',') if x.isdigit()]
            articles = (
                Article.visible()
                .filter(Article.id.in_(ids_list))
                .options(
                    joinedload(Article.issue).joinedload(Issue.journal),
                    joinedload(Article.article_authors)
                )
                .all()
            )
        else:
            articles = _filtered_articles_query(journal_id, status_filter, search).all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['№', 'Название', 'Авторы', 'Журнал', 'Выпуск', 'Дата поступления', 'Оплата', 'Рецензия', 'Редакт.', 'Акт эксп.'])

        for idx, article in enumerate(articles, 1):
            authors = ", ".join([a.full_name for a in article.article_authors]) if article.article_authors else (article.authors or "-")
            journal_name = article.issue.journal.name if article.issue and article.issue.journal else "-"
            issue_info = f"№{article.issue.number}/{article.issue.year}" if article.issue else "-"

            writer.writerow([
                idx,
                article.title or "-",
                authors,
                journal_name,
                issue_info,
                article.submission_date or "-",
                "Да" if article.payment_received else "Нет",
                "Да" if article.has_review else "Нет",
                "Да" if article.edited else "Нет",
                "Да" if article.has_expertise_act else "Нет",
            ])

        output.seek(0)
        bytes_output = BytesIO(output.getvalue().encode('utf-8-sig'))

        filename = f"articles_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return send_file(bytes_output, mimetype='text/csv; charset=utf-8',
                         as_attachment=True, download_name=filename)

    # ==================== USER MANAGEMENT ====================
    @app.route('/admin/users')
    @admin_required
    def admin_users():
        users = User.query.order_by(User.role.desc(), User.display_name).all()
        return render_template('admin/users.html', users=users)

    @app.route('/admin/users/add', methods=['POST'])
    @admin_required
    def admin_user_add():
        username = request.form.get('username', '').strip()
        display_name = request.form.get('display_name', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'user')

        if not username or not display_name or not password:
            return jsonify({'success': False, 'error': 'Заполните все поля'}), 400

        if User.query.filter_by(username=username).first():
            return jsonify({'success': False, 'error': 'Пользователь с таким логином уже существует'}), 400

        if role not in ('admin', 'user'):
            role = 'user'

        user = User(username=username, display_name=display_name, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.flush()  # получаем user.id для лога

        log_activity('created', 'user', user.id, display_name, f'Роль: {role}')
        db.session.commit()

        return jsonify({'success': True, 'user': {
            'id': user.id, 'username': user.username,
            'display_name': user.display_name, 'role': user.role
        }})

    @app.route('/admin/users/<int:user_id>/update', methods=['POST'])
    @admin_required
    def admin_user_update(user_id):
        user = User.query.get_or_404(user_id)
        display_name = request.form.get('display_name', '').strip()
        role = request.form.get('role', 'user')
        password = request.form.get('password', '').strip()
        is_active = request.form.get('is_active', 'true') == 'true'

        if not display_name:
            return jsonify({'success': False, 'error': 'Имя не может быть пустым'}), 400

        if role not in ('admin', 'user'):
            role = 'user'

        # Не дать лишить себя прав
        if user.id == current_user.id and role != 'admin':
            return jsonify({'success': False, 'error': 'Вы не можете снять с себя роль администратора'}), 400

        user.display_name = display_name
        user.role = role
        user.is_active_user = is_active

        if password:
            user.set_password(password)

        log_activity('updated', 'user', user.id, display_name, f'Роль: {role}')
        db.session.commit()

        return jsonify({'success': True})

    @app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
    @admin_required
    def admin_user_delete(user_id):
        user = User.query.get_or_404(user_id)

        if user.id == current_user.id:
            return jsonify({'success': False, 'error': 'Вы не можете удалить самого себя'}), 400

        name = user.display_name
        log_activity('deleted', 'user', user_id, name)
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True})

    # =============================================
    #   КОРЗИНА (восстановление удалённых)
    # =============================================
    @app.route('/admin/trash')
    @admin_required
    def admin_trash():
        """Список удалённых статей и выпусков с возможностью восстановления."""
        deleted_articles = (
            Article.query
            .filter(Article.deleted_at.isnot(None))
            .options(
                joinedload(Article.issue).joinedload(Issue.journal),
                joinedload(Article.article_authors)
            )
            .order_by(Article.deleted_at.desc())
            .all()
        )
        deleted_issues = (
            Issue.query
            .filter(Issue.deleted_at.isnot(None))
            .options(joinedload(Issue.journal))
            .order_by(Issue.deleted_at.desc())
            .all()
        )
        return render_template(
            'admin/trash.html',
            deleted_articles=deleted_articles,
            deleted_issues=deleted_issues,
        )

    @app.route('/admin/trash/article/<int:article_id>/restore', methods=['POST'])
    @admin_required
    def admin_trash_restore_article(article_id):
        article = Article.query.filter(Article.id == article_id, Article.deleted_at.isnot(None)).first_or_404()
        article.deleted_at = None
        db.session.commit()
        log_activity('restored', 'article', article.id, article.title)
        flash('Статья восстановлена', 'success')
        return redirect(url_for('admin_trash'))

    @app.route('/admin/trash/issue/<int:issue_id>/restore', methods=['POST'])
    @admin_required
    def admin_trash_restore_issue(issue_id):
        issue = Issue.query.filter(Issue.id == issue_id, Issue.deleted_at.isnot(None)).first_or_404()
        issue.deleted_at = None
        for article in issue.articles:
            article.deleted_at = None
        db.session.commit()
        log_activity('restored', 'issue', issue.id, f'№{issue.number}/{issue.year}', f'Журнал: {issue.journal.name}')
        flash(f'Выпуск №{issue.number}/{issue.year} и его статьи восстановлены', 'success')
        return redirect(url_for('admin_trash'))

    # =============================================
    #   СТРАНИЦА ЭКСПОРТА
    # =============================================
    @app.route('/admin/export')
    @admin_required
    def admin_export():
        """Страница экспорта с выбором формата и фильтров."""
        journals = Journal.query.order_by(Journal.name).all()
        total_articles = Article.visible().count()
        return render_template('admin/export.html', journals=journals, total_articles=total_articles)

    # =============================================
    #   БЭКАПЫ
    # =============================================
    @app.route('/admin/backups')
    @admin_required
    def admin_backups():
        base_dir = os.path.dirname(os.path.abspath(__file__))
        backups_dir = os.path.join(base_dir, 'backups')
        os.makedirs(backups_dir, exist_ok=True)

        backups = []
        for f in sorted(os.listdir(backups_dir), reverse=True):
            if f.endswith('.db'):
                path = os.path.join(backups_dir, f)
                size = os.path.getsize(path)
                mtime = datetime.fromtimestamp(os.path.getmtime(path))
                backups.append({'name': f, 'size': size, 'date': mtime})

        db_path = os.path.join(base_dir, 'instance', 'journal.db')
        db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0

        return render_template('admin/backups.html', backups=backups, db_size=db_size)

    @app.route('/admin/backups/create', methods=['POST'])
    @admin_required
    def admin_backup_create():
        from backup import create_backup
        create_backup()
        return jsonify({'success': True, 'message': 'Бэкап создан'})

    @app.route('/admin/backups/download/<filename>')
    @admin_required
    def admin_backup_download(filename):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        backups_dir = os.path.join(base_dir, 'backups')
        safe = secure_filename(filename)
        path = os.path.join(backups_dir, safe)
        if not os.path.exists(path):
            return 'Файл не найден', 404
        return send_file(path, as_attachment=True, download_name=safe)

    @app.route('/admin/backups/download-current')
    @admin_required
    def admin_backup_download_current():
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, 'instance', 'journal.db')
        if not os.path.exists(db_path):
            return 'БД не найдена', 404
        name = f'journal_current_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.db'
        return send_file(db_path, as_attachment=True, download_name=name)

    @app.route('/admin/backups/restore', methods=['POST'])
    @admin_required
    def admin_backup_restore():
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, 'instance', 'journal.db')
        backups_dir = os.path.join(base_dir, 'backups')

        # Сначала создаём бэкап текущей БД
        from backup import create_backup
        create_backup()

        source = request.form.get('backup_name')
        uploaded = request.files.get('backup_file')

        if uploaded and uploaded.filename:
            # Восстановление из загруженного файла
            db.session.remove()
            db.engine.dispose()
            uploaded.save(db_path)
            return jsonify({'success': True, 'message': 'БД восстановлена из загруженного файла. Перезагрузите приложение.'})
        elif source:
            # Восстановление из существующего бэкапа
            safe = secure_filename(source)
            src_path = os.path.join(backups_dir, safe)
            if not os.path.exists(src_path):
                return jsonify({'success': False, 'error': 'Бэкап не найден'}), 404
            db.session.remove()
            db.engine.dispose()
            import shutil
            shutil.copy(src_path, db_path)
            return jsonify({'success': True, 'message': f'БД восстановлена из {safe}. Перезагрузите приложение.'})
        else:
            return jsonify({'success': False, 'error': 'Не указан источник'}), 400

    @app.route('/admin/backups/delete/<filename>', methods=['POST'])
    @admin_required
    def admin_backup_delete(filename):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        backups_dir = os.path.join(base_dir, 'backups')
        safe = secure_filename(filename)
        path = os.path.join(backups_dir, safe)
        if os.path.exists(path):
            os.remove(path)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Файл не найден'}), 404

    # =============================================
    #   ПАРСИНГ PDF (авто-заполнение из статьи)
    # =============================================
    @app.route('/api/parse-pdf', methods=['POST'])
    @login_required
    def parse_pdf():
        """Извлекает название и авторов из загруженного PDF или Word-документа."""
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'error': 'Файл не выбран'}), 400

        fname_lower = file.filename.lower()
        allowed_ext = ('.pdf', '.docx', '.doc')
        if not fname_lower.endswith(allowed_ext):
            return jsonify({'error': 'Поддерживаются форматы: PDF, DOCX, DOC'}), 400

        # Сохраняем временно для парсинга
        upload_folder = current_app.config['UPLOAD_FOLDER']
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
        filename = secure_filename(timestamp + file.filename)
        filepath = os.path.join(upload_folder, filename)
        file.save(filepath)

        try:
            if fname_lower.endswith('.pdf'):
                from pdf_parser import parse_article_pdf
                result = parse_article_pdf(filepath)
            else:
                from pdf_parser import parse_article_docx
                result = parse_article_docx(filepath)
            result['saved_filename'] = filename
            return jsonify(result)
        except ImportError as e:
            return jsonify({'error': f'Отсутствует модуль: {str(e)}'}), 500
        except Exception as e:
            return jsonify({'error': f'Ошибка парсинга: {str(e)}'}), 500

    # =============================================
    #   ИСТОРИЯ ИЗМЕНЕНИЙ СТАТЬИ
    # =============================================
    @app.route('/article/<int:article_id>/history')
    @login_required
    def article_history(article_id):
        article = Article.visible().filter_by(id=article_id).first_or_404()
        entries = ArticleHistory.query.filter_by(article_id=article_id).order_by(ArticleHistory.created_at.desc()).limit(50).all()
        result = []
        for e in entries:
            result.append({
                'user': e.user_name,
                'action': e.action,
                'changes': json.loads(e.changes) if e.changes else [],
                'date': e.created_at.strftime('%d.%m.%Y %H:%M')
            })
        return jsonify({'article': article.title, 'history': result})

    # =============================================
    #   QUICK VIEW — JSON-данные статьи
    # =============================================
    @app.route('/article/<int:article_id>/json')
    @login_required
    def article_json(article_id):
        article = Article.visible().options(
            joinedload(Article.article_authors),
            joinedload(Article.images),
            joinedload(Article.issue).joinedload(Issue.journal)
        ).filter_by(id=article_id).first_or_404()

        authors = []
        for a in sorted(article.article_authors, key=lambda x: x.order):
            authors.append({
                'name': a.full_name,
                'email': a.email or '',
                'organization': a.organization or '',
                'degree': a.degree or '',
                'position': a.position or '',
                'phone': a.phone or ''
            })

        images = [{'filename': img.filename} for img in sorted(article.images, key=lambda x: x.order)]

        issue = article.issue
        journal = issue.journal if issue else None

        comment_count = ArticleComment.query.filter_by(article_id=article_id).count()

        return jsonify({
            'id': article.id,
            'title': article.title,
            'authors': authors,
            'authors_str': article.authors or '',
            'submission_date': article.submission_date or '',
            'payment': article.payment_received,
            'review': article.has_review,
            'edited': article.edited,
            'expertise': article.has_expertise_act,
            'expertise_act_file': article.expertise_act_file or '',
            'notes': article.notes or '',
            'manuscript_file': article.manuscript_file or '',
            'review_file': article.review_file or '',
            'title_pdf': article.title_pdf or '',
            'images': images,
            'issue': f'№{issue.number}/{issue.year}' if issue else '',
            'journal': journal.name if journal else '',
            'comment_count': comment_count
        })

    # =============================================
    #   EDIT-DATA — JSON-данные для панели редактирования
    # =============================================
    @app.route('/article/<int:article_id>/edit-data')
    @login_required
    def article_edit_data(article_id):
        article = Article.visible().options(
            joinedload(Article.article_authors),
            joinedload(Article.images),
            joinedload(Article.issue).joinedload(Issue.journal)
        ).filter_by(id=article_id).first_or_404()

        authors = []
        for a in sorted(article.article_authors, key=lambda x: x.order):
            authors.append({
                'name': a.full_name,
                'email': a.email or '',
                'organization': a.organization or '',
                'degree': a.degree or '',
                'position': a.position or '',
                'phone': a.phone or ''
            })

        images = [{'id': img.id, 'filename': img.filename} for img in sorted(article.images, key=lambda x: x.order)]

        # Разбиваем дату
        day, month, year = '', '', ''
        if article.submission_date:
            parts = article.submission_date.split('.')
            if len(parts) == 3:
                day, month, year = parts

        issue = article.issue
        journal = issue.journal if issue else None

        return jsonify({
            'id': article.id,
            'title': article.title or '',
            'authors': authors,
            'submission_day': day,
            'submission_month': month,
            'submission_year': year,
            'notes': article.notes or '',
            'manuscript_file': article.manuscript_file or '',
            'review_file': article.review_file or '',
            'title_pdf': article.title_pdf or '',
            'expertise_act_file': article.expertise_act_file or '',
            'images': images,
            'issue_id': article.issue_id,
            'journal_name': journal.name if journal else '',
            'issue_label': f'№{issue.number}/{issue.year}' if issue else '',
        })

    # =============================================
    #   КОММЕНТАРИИ К СТАТЬЕ
    # =============================================
    @app.route('/article/<int:article_id>/comments')
    @login_required
    def article_comments(article_id):
        comments = ArticleComment.query.filter_by(article_id=article_id).order_by(ArticleComment.created_at.asc()).all()
        return jsonify([{
            'id': c.id,
            'user': c.user_name,
            'text': c.text,
            'date': c.created_at.strftime('%d.%m.%Y %H:%M')
        } for c in comments])

    @app.route('/article/<int:article_id>/comments', methods=['POST'])
    @login_required
    def add_article_comment(article_id):
        Article.visible().filter_by(id=article_id).first_or_404()
        data = request.get_json()
        text = (data.get('text') or '').strip()
        if not text:
            return jsonify({'success': False, 'error': 'Пустой комментарий'}), 400

        comment = ArticleComment(
            article_id=article_id,
            user_name=current_user.display_name,
            text=text
        )
        db.session.add(comment)
        db.session.commit()
        return jsonify({
            'success': True,
            'comment': {
                'id': comment.id,
                'user': comment.user_name,
                'text': comment.text,
                'date': comment.created_at.strftime('%d.%m.%Y %H:%M')
            }
        })

    # ─── Авторы ────────────────────────────────────────────────────────

    def _normalize_author_name(name):
        """
        Нормализует имя автора для группировки.
        'В.В. Беляев'  -> 'беляев_вв'
        'Беляев В.В.'  -> 'беляев_вв'
        'Владимир Владимирович Беляев' -> 'беляев_вв'
        'Беляев Владимир Владимирович' -> 'беляев_вв'
        """
        if not name:
            return ''
        name = name.strip()
        # Убираем цифры-индексы, скобки и лишние символы
        name = re.sub(r'[\d*†‡§]+', '', name)
        name = re.sub(r'[()«»""\']+', '', name)
        name = name.strip(' ,;.')

        # Разбиваем на части
        parts = name.split()
        if not parts:
            return ''

        # Определяем фамилию и инициалы
        # Инициал — часть из 1-2 букв (возможно с точкой): "В.", "В.В.", "А"
        initials = []
        surname = ''
        for p in parts:
            clean = p.rstrip('.')
            # Инициал: одна буква, или две слитные буквы (напр. "А.В" или "АВ")
            if len(clean) <= 2 and clean.isalpha():
                initials.append(clean[0].lower())
            elif '.' in p and all(len(x) <= 1 for x in p.split('.') if x):
                # "А.В." -> ['А', 'В']
                for x in p.split('.'):
                    if x and len(x) == 1:
                        initials.append(x[0].lower())
            elif len(clean) > 2:
                # Полное слово — кандидат на фамилию или имя/отчество
                if not surname:
                    surname = clean
                else:
                    # Фамилия — обычно самое короткое из длинных слов (не всегда),
                    # но надёжнее: если уже есть фамилия, это имя/отчество → берём первую букву
                    initials.append(clean[0].lower())

        # Если фамилия не определена (все части были инициалами)
        if not surname and parts:
            surname = parts[-1].rstrip('.')

        key = surname.lower().rstrip('.')
        if initials:
            key += '_' + ''.join(sorted(initials))

        return key

    def _group_authors():
        """
        Загружает всех ArticleAuthor и группирует их.
        Приоритет 1: по email (одинаковый email = один автор).
        Приоритет 2: по нормализованному имени.
        Возвращает список dict (отсортирован по фамилии):
            {name, email, organization, articles: [{id, title, journal, issue}], name_variants}
        """
        all_authors = (
            ArticleAuthor.query
            .options(
                joinedload(ArticleAuthor.article)
                .joinedload(Article.issue)
                .joinedload(Issue.journal)
            )
            .all()
        )

        # Шаг 1: группируем по email (если есть)
        email_groups = defaultdict(list)     # email -> [ArticleAuthor]
        no_email = []                        # авторы без email
        for aa in all_authors:
            email = (aa.email or '').strip().lower()
            if email:
                email_groups[email].append(aa)
            else:
                no_email.append(aa)

        # Шаг 2: авторов без email группируем по нормализованному имени
        name_groups = defaultdict(list)
        for aa in no_email:
            key = _normalize_author_name(aa.full_name)
            if key:
                name_groups[key].append(aa)
            else:
                name_groups[f'_unknown_{aa.id}'] = [aa]

        # Шаг 3: объединяем email-группы в финальный результат
        result = []

        def _build_entry(members):
            """Строит запись автора из группы ArticleAuthor."""
            names = set()
            emails = set()
            orgs = set()
            articles = []
            seen_articles = set()

            for aa in members:
                if aa.full_name:
                    names.add(aa.full_name.strip())
                if aa.email and aa.email.strip():
                    emails.add(aa.email.strip().lower())
                if aa.organization and aa.organization.strip():
                    orgs.add(aa.organization.strip())
                if aa.article and aa.article_id not in seen_articles:
                    seen_articles.add(aa.article_id)
                    art = aa.article
                    journal_name = ''
                    issue_label = ''
                    if art.issue:
                        if art.issue.journal:
                            journal_name = art.issue.journal.name
                        issue_label = f'\u2116{art.issue.number}/{art.issue.year}'
                    articles.append({
                        'id': art.id,
                        'title': art.title or '',
                        'journal': journal_name,
                        'issue': issue_label,
                    })

            # Основное имя — самое длинное
            name_list = sorted(names, key=len, reverse=True)
            main_name = name_list[0] if name_list else '—'
            variants = [n for n in name_list[1:] if n != main_name]

            return {
                'name': main_name,
                'name_variants': variants,
                'email': ', '.join(sorted(emails)),
                'organization': '; '.join(sorted(orgs, key=len, reverse=True)[:2]),
                'article_count': len(articles),
                'articles': sorted(articles, key=lambda x: x['id'], reverse=True),
            }

        for email, members in email_groups.items():
            result.append(_build_entry(members))

        for key, members in name_groups.items():
            result.append(_build_entry(members))

        # Сортируем по фамилии (первое слово)
        result.sort(key=lambda x: x['name'].split()[-1].lower() if x['name'] else '')

        return result

    @app.route('/admin/authors')
    @admin_required
    def admin_authors():
        """Страница со списком всех авторов — данные через AJAX."""
        search = request.args.get('search', '').strip()
        total_authors = db.session.query(db.func.count(db.distinct(ArticleAuthor.full_name))).scalar()
        return render_template('admin/authors.html',
                               total=total_authors,
                               search=search)

    @app.route('/admin/authors/api')
    @admin_required
    def admin_authors_api():
        """JSON API для пагинированного списка авторов (группировка на уровне SQL)."""
        search = request.args.get('search', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 50, type=int), 200)
        sort = request.args.get('sort', 'name_asc')  # name_asc, name_desc, articles_desc, articles_asc, email

        # SQL-группировка авторов по LOWER(full_name)
        article_count_col = db.func.count(db.distinct(ArticleAuthor.article_id)).label('article_count')
        base_q = (
            db.session.query(
                ArticleAuthor.full_name,
                db.func.min(ArticleAuthor.email).label('email'),
                db.func.min(ArticleAuthor.organization).label('organization'),
                article_count_col,
            )
            .filter(ArticleAuthor.full_name.isnot(None))
            .filter(ArticleAuthor.full_name != '')
            .filter(~ArticleAuthor.full_name.like('(%'))   # организации/города
            .filter(~ArticleAuthor.full_name.like('-%'))   # фрагменты текста
            .filter(~ArticleAuthor.full_name.like('<%'))   # HTML-теги
            .filter(~ArticleAuthor.full_name.like('%<%'))   # HTML внутри имени
            .filter(db.func.length(ArticleAuthor.full_name) > 2)  # слишком короткие
        )

        if search:
            like = f'%{search}%'
            base_q = base_q.filter(
                db.or_(
                    ArticleAuthor.full_name.ilike(like),
                    ArticleAuthor.email.ilike(like),
                    ArticleAuthor.organization.ilike(like),
                )
            )

        base_q = base_q.group_by(db.func.lower(ArticleAuthor.full_name))

        # Сортировка
        sort_map = {
            'name_asc': ArticleAuthor.full_name.asc(),
            'name_desc': ArticleAuthor.full_name.desc(),
            'articles_desc': db.desc('article_count'),
            'articles_asc': db.asc('article_count'),
            'email': db.func.min(ArticleAuthor.email).asc(),
        }
        order = sort_map.get(sort, ArticleAuthor.full_name.asc())

        total = base_q.count()
        rows = (
            base_q
            .order_by(order)
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        result = []
        for row in rows:
            result.append({
                'name': row.full_name or '',
                'email': (row.email or '').strip(),
                'organization': (row.organization or '').strip(),
                'article_count': row.article_count,
            })

        total_pages = (total + per_page - 1) // per_page
        return jsonify({
            'authors': result,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
        })

    @app.route('/admin/authors/articles')
    @admin_required
    def admin_author_articles():
        """JSON API — статьи конкретного автора (по имени)."""
        name = request.args.get('name', '').strip()
        if not name:
            return jsonify({'articles': []})

        # SQLite lower() переопределён через Python для поддержки кириллицы
        authors = (
            ArticleAuthor.query
            .filter(db.func.lower(ArticleAuthor.full_name) == name.lower())
            .options(
                joinedload(ArticleAuthor.article)
                .joinedload(Article.issue)
                .joinedload(Issue.journal)
            )
            .all()
        )

        seen = set()
        articles = []
        for aa in authors:
            if aa.article and aa.article_id not in seen:
                seen.add(aa.article_id)
                art = aa.article
                journal_name = ''
                issue_label = ''
                if art.issue:
                    if art.issue.journal:
                        journal_name = art.issue.journal.name
                    issue_label = f'\u2116{art.issue.number}/{art.issue.year}'
                articles.append({
                    'id': art.id,
                    'title': art.title or '',
                    'journal': journal_name,
                    'issue': issue_label,
                })

        articles.sort(key=lambda x: x['id'], reverse=True)
        return jsonify({'articles': articles})
