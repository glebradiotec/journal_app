"""
Общий модуль генерации Excel-файлов для экспорта статей.
Используется в routes_admin.py (веб-экспорт) и telegram_bot.py (Telegram-бот).
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# === Стили ===
_THIN_BORDER = Border(
    left=Side(style='thin', color='D0D7DE'),
    right=Side(style='thin', color='D0D7DE'),
    top=Side(style='thin', color='D0D7DE'),
    bottom=Side(style='thin', color='D0D7DE'),
)
_TITLE_FONT = Font(bold=True, size=14, color='24292F')
_SUBTITLE_FONT = Font(size=11, color='57606A')
_HEADER_FILL = PatternFill(start_color='2D333B', end_color='2D333B', fill_type='solid')
_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_ROW_EVEN_FILL = PatternFill(start_color='F6F8FA', end_color='F6F8FA', fill_type='solid')
_STATUS_YES_FILL = PatternFill(start_color='DAFBE1', end_color='DAFBE1', fill_type='solid')
_STATUS_YES_FONT = Font(bold=True, color='1A7F37', size=11)
_STATUS_NO_FILL = PatternFill(start_color='FFEBE9', end_color='FFEBE9', fill_type='solid')
_STATUS_NO_FONT = Font(bold=True, color='CF222E', size=11)
_SUMMARY_FILL = PatternFill(start_color='DDF4FF', end_color='DDF4FF', fill_type='solid')
_SUMMARY_FONT = Font(bold=True, size=11, color='0969DA')

_COL_WIDTHS = {'A': 5, 'B': 42, 'C': 32, 'D': 22, 'E': 13, 'F': 17, 'G': 11, 'H': 11, 'I': 11, 'J': 12}
_HEADERS = ['№', 'Название', 'Авторы', 'Журнал', 'Выпуск', 'Дата поступления', 'Оплата', 'Рецензия', 'Редакт.', 'Акт эксп.']


def generate_articles_excel(rows, filter_desc='Все статьи'):
    """
    Генерирует Excel-файл со статьями.

    Args:
        rows: список словарей, каждый содержит:
            - title, authors, journal_name, issue_info, submission_date
            - payment (bool), review (bool), edited (bool), expertise (bool)
        filter_desc: текстовое описание фильтра для подзаголовка

    Returns:
        BytesIO с готовым .xlsx файлом и количество строк (bytes_io, total)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Статьи"

    # --- Заголовок отчёта ---
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'Экспорт статей — Radiotec Journal App'
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    sub_cell = ws['A2']
    sub_cell.value = f'{filter_desc}  |  Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    sub_cell.font = _SUBTITLE_FONT
    sub_cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    # --- Заголовки таблицы (строка 4) ---
    for col_idx, h in enumerate(_HEADERS, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[4].height = 26

    # --- Данные ---
    paid_count = reviewed_count = edited_count = expertise_count = 0

    for row_idx, row in enumerate(rows, 1):
        r = row_idx + 4

        is_paid = bool(row.get('payment', False))
        is_reviewed = bool(row.get('review', False))
        is_edited = bool(row.get('edited', False))
        has_expertise = bool(row.get('expertise', False))

        if is_paid: paid_count += 1
        if is_reviewed: reviewed_count += 1
        if is_edited: edited_count += 1
        if has_expertise: expertise_count += 1

        row_data = [
            row_idx,
            row.get('title', '-') or '-',
            row.get('authors', '-') or '-',
            row.get('journal_name', '-') or '-',
            row.get('issue_info', '-') or '-',
            row.get('submission_date', '-') or '-',
            'Да' if is_paid else 'Нет',
            'Да' if is_reviewed else 'Нет',
            'Да' if is_edited else 'Нет',
            'Да' if has_expertise else 'Нет',
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 2))
            if row_idx % 2 == 0:
                cell.fill = _ROW_EVEN_FILL

        # Условное форматирование статусов (колонки 7-10)
        for col in (7, 8, 9, 10):
            cell = ws.cell(row=r, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value == 'Да':
                cell.fill = _STATUS_YES_FILL
                cell.font = _STATUS_YES_FONT
            else:
                cell.fill = _STATUS_NO_FILL
                cell.font = _STATUS_NO_FONT

    total = len(rows)

    # --- Итоговая строка ---
    if total > 0:
        sr = total + 5
        ws.row_dimensions[sr].height = 26
        ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=2)
        summary_cell = ws.cell(row=sr, column=1, value=f'Итого: {total} статей')
        summary_cell.font = _SUMMARY_FONT
        summary_cell.fill = _SUMMARY_FILL
        summary_cell.alignment = Alignment(vertical='center')
        summary_cell.border = _THIN_BORDER
        ws.cell(row=sr, column=2).fill = _SUMMARY_FILL
        ws.cell(row=sr, column=2).border = _THIN_BORDER

        for c in range(3, 7):
            cell = ws.cell(row=sr, column=c, value='')
            cell.fill = _SUMMARY_FILL
            cell.border = _THIN_BORDER

        for col, count in [(7, paid_count), (8, reviewed_count), (9, edited_count), (10, expertise_count)]:
            cell = ws.cell(row=sr, column=col, value=f'{count}/{total}')
            cell.font = _SUMMARY_FONT
            cell.fill = _SUMMARY_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = _THIN_BORDER

    # --- Ширина колонок ---
    for col_letter, w in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = 'A5'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, total
